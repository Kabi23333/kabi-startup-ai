"""
generer_kundeliste.py
Generates a list of 100 potential Norwegian small business customers for KABI AUTOMATION.
Saves to kundeliste.xlsx and outreach-liste.md
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os

# ── 1. DATA ──────────────────────────────────────────────────────────────────

kunder = [
    # HÅNDVERKERE (30)
    {"Nr": 1,  "Bedriftsnavn": "Hansen Rør & Varme AS",        "Bransje": "Rørlegger",                    "By": "Oslo",         "Manuelt problem": "Manuell fakturaregistrering i Excel",                       "KABI-løsning": "Excel-automatisering + rapport",       "Pris (NOK)": "6000-8000",   "Finn dem på": "Google Maps / Finn.no"},
    {"Nr": 2,  "Bedriftsnavn": "Nordvik Elektro",               "Bransje": "Elektriker",                   "By": "Bergen",       "Manuelt problem": "Tidsskjema og timelister på papir",                         "KABI-løsning": "Excel-automatisering",                 "Pris (NOK)": "5000-7000",   "Finn dem på": "Facebook Bedrifter"},
    {"Nr": 3,  "Bedriftsnavn": "Bakke Snekker & Bygg",          "Bransje": "Snekker",                      "By": "Trondheim",    "Manuelt problem": "Tilbudsskriving tar 2-3 timer per kunde",                   "KABI-løsning": "Email-automasjon + mal",               "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 4,  "Bedriftsnavn": "Solberg VVS",                   "Bransje": "Rørlegger",                    "By": "Stavanger",    "Manuelt problem": "Manuelle purringer til kunder på ubetalte fakturaer",       "KABI-løsning": "Email-automasjon",                     "Pris (NOK)": "5000-6000",   "Finn dem på": "Google Maps"},
    {"Nr": 5,  "Bedriftsnavn": "Østlund Maler AS",              "Bransje": "Maler",                        "By": "Oslo",         "Manuelt problem": "Rapportskriving etter hvert oppdrag",                       "KABI-løsning": "Rapport-generator",                    "Pris (NOK)": "4000-5000",   "Finn dem på": "Finn.no"},
    {"Nr": 6,  "Bedriftsnavn": "Fjord Elektrikere",             "Bransje": "Elektriker",                   "By": "Ålesund",      "Manuelt problem": "Bestilling av materialer per telefon og e-post manuelt",    "KABI-løsning": "Excel-automatisering",                 "Pris (NOK)": "5000-7000",   "Finn dem på": "Google Maps"},
    {"Nr": 7,  "Bedriftsnavn": "Strand Tømrer",                 "Bransje": "Snekker",                      "By": "Kristiansand", "Manuelt problem": "Holder styr på materiallager i hodet",                      "KABI-løsning": "Lager-tracker i Excel",                "Pris (NOK)": "4000-6000",   "Finn dem på": "Facebook"},
    {"Nr": 8,  "Bedriftsnavn": "Viken Rørservice",              "Bransje": "Rørlegger",                    "By": "Drammen",      "Manuelt problem": "Kundeoppfølging gjøres ikke systematisk",                   "KABI-løsning": "CRM i Excel + email-mal",              "Pris (NOK)": "6000-8000",   "Finn dem på": "Google Maps"},
    {"Nr": 9,  "Bedriftsnavn": "Haugen Byggservice",            "Bransje": "Snekker",                      "By": "Tromsø",       "Manuelt problem": "Fakturering skjer for sent",                                "KABI-løsning": "Faktura-påminner i Excel",             "Pris (NOK)": "4000-5000",   "Finn dem på": "Facebook Bedrifter"},
    {"Nr": 10, "Bedriftsnavn": "Lunde Elektro AS",              "Bransje": "Elektriker",                   "By": "Fredrikstad",  "Manuelt problem": "HMS-rapporter skrives manuelt",                             "KABI-løsning": "Rapport-generator",                    "Pris (NOK)": "5000-7000",   "Finn dem på": "LinkedIn"},
    {"Nr": 11, "Bedriftsnavn": "Aas & Sønner Rør",             "Bransje": "Rørlegger",                    "By": "Sandnes",      "Manuelt problem": "Manglende oversikt over aktive jobber",                     "KABI-løsning": "Jobb-tracker Excel",                   "Pris (NOK)": "5000-6000",   "Finn dem på": "Google Maps"},
    {"Nr": 12, "Bedriftsnavn": "Berg Maler og Fasade",          "Bransje": "Maler",                        "By": "Bodø",         "Manuelt problem": "Kundekommunikasjon via SMS og lapper",                      "KABI-løsning": "Email-automasjon",                     "Pris (NOK)": "4000-5000",   "Finn dem på": "Facebook"},
    {"Nr": 13, "Bedriftsnavn": "Dahl VVS og Baderom",           "Bransje": "Rørlegger",                    "By": "Tønsberg",     "Manuelt problem": "Kalkulasjon av timepris gjøres feil",                       "KABI-løsning": "Kalkulator-verktøy Excel",             "Pris (NOK)": "6000-8000",   "Finn dem på": "Finn.no"},
    {"Nr": 14, "Bedriftsnavn": "Nilsen Elektriske",             "Bransje": "Elektriker",                   "By": "Moss",         "Manuelt problem": "Servicehistorikk på kunder føres ikke",                     "KABI-løsning": "Historikk-database Excel",             "Pris (NOK)": "5000-7000",   "Finn dem på": "Google Maps"},
    {"Nr": 15, "Bedriftsnavn": "Kvam Bygg & Renovering",        "Bransje": "Snekker",                      "By": "Hamar",        "Manuelt problem": "Alle avtaler koordineres per telefon",                      "KABI-løsning": "Koordinerings-Excel + mal",            "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook"},
    {"Nr": 16, "Bedriftsnavn": "Rørvik Sanitær",                "Bransje": "Rørlegger",                    "By": "Larvik",       "Manuelt problem": "Faktura lages i Word, ingen system",                        "KABI-løsning": "Excel-fakturering",                    "Pris (NOK)": "4000-5000",   "Finn dem på": "Google Maps"},
    {"Nr": 17, "Bedriftsnavn": "Lie Elektro og Sikkerhet",      "Bransje": "Elektriker",                   "By": "Sarpsborg",    "Manuelt problem": "Manuelle sjekklister for hvert oppdrag",                    "KABI-løsning": "Sjekkliste-generator",                 "Pris (NOK)": "5000-6000",   "Finn dem på": "LinkedIn"},
    {"Nr": 18, "Bedriftsnavn": "Munthe Tømrer AS",              "Bransje": "Snekker",                      "By": "Gjøvik",       "Manuelt problem": "Timeregistrering på papir",                                 "KABI-løsning": "Excel timesheet + rapport",            "Pris (NOK)": "4000-5000",   "Finn dem på": "Facebook Bedrifter"},
    {"Nr": 19, "Bedriftsnavn": "Holm Maling og Sparkling",      "Bransje": "Maler",                        "By": "Skien",        "Manuelt problem": "Ingen oversikt over hvilke kunder som er lønnsomme",        "KABI-løsning": "Lønnsomhets-analyse Excel",            "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook"},
    {"Nr": 20, "Bedriftsnavn": "Elgaard Rørlegging",            "Bransje": "Rørlegger",                    "By": "Porsgrunn",    "Manuelt problem": "Tilbud sendes ikke ut raskt nok",                           "KABI-løsning": "Email-mal for hurtigtilbud",           "Pris (NOK)": "5000-6000",   "Finn dem på": "Google Maps"},
    {"Nr": 21, "Bedriftsnavn": "Thorsen Elektro",               "Bransje": "Elektriker",                   "By": "Arendal",      "Manuelt problem": "Lønn til ansatte beregnes manuelt",                         "KABI-løsning": "Excel lønnskalkulator",                "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 22, "Bedriftsnavn": "Berge Byggtjenester",           "Bransje": "Snekker",                      "By": "Lillehammer",  "Manuelt problem": "Ingen system for å følge opp leads",                        "KABI-løsning": "Lead-tracker Excel",                   "Pris (NOK)": "5000-7000",   "Finn dem på": "Facebook"},
    {"Nr": 23, "Bedriftsnavn": "Engen VVS Service",             "Bransje": "Rørlegger",                    "By": "Molde",        "Manuelt problem": "Garantioppfølging glemmes",                                 "KABI-løsning": "Garanti-tracker Excel",                "Pris (NOK)": "4000-5000",   "Finn dem på": "Google Maps"},
    {"Nr": 24, "Bedriftsnavn": "Myhre Elektrikeren",            "Bransje": "Elektriker",                   "By": "Harstad",      "Manuelt problem": "Timeforbruk per prosjekt ukjent",                           "KABI-løsning": "Prosjekt-timer Excel",                 "Pris (NOK)": "5000-6000",   "Finn dem på": "Facebook Bedrifter"},
    {"Nr": 25, "Bedriftsnavn": "Solli Snekker & Interiør",      "Bransje": "Snekker",                      "By": "Kongsberg",    "Manuelt problem": "Ingen standardiserte tilbudsmaler",                         "KABI-løsning": "Tilbuds-generator",                    "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 26, "Bedriftsnavn": "Birkeland Rør",                 "Bransje": "Rørlegger",                    "By": "Halden",       "Manuelt problem": "SMS-varsling til kunder gjøres manuelt",                    "KABI-løsning": "Email-automasjon",                     "Pris (NOK)": "4000-5000",   "Finn dem på": "Google Maps"},
    {"Nr": 27, "Bedriftsnavn": "Finstad Maler",                 "Bransje": "Maler",                        "By": "Horten",       "Manuelt problem": "Ingen digital registrering av jobber",                      "KABI-løsning": "Jobb-logg Excel",                      "Pris (NOK)": "4000-5000",   "Finn dem på": "Facebook"},
    {"Nr": 28, "Bedriftsnavn": "Løken Elektro AS",              "Bransje": "Elektriker",                   "By": "Steinkjer",    "Manuelt problem": "Rapporter til forsikring skrives manuelt",                  "KABI-løsning": "Rapport-generator",                    "Pris (NOK)": "5000-7000",   "Finn dem på": "LinkedIn"},
    {"Nr": 29, "Bedriftsnavn": "Wold Bygg & Rehabilitering",    "Bransje": "Snekker",                      "By": "Ålesund",      "Manuelt problem": "Prosjektbudsjett sprekker uten varsel",                     "KABI-løsning": "Budsjett-tracker Excel",               "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook Bedrifter"},
    {"Nr": 30, "Bedriftsnavn": "Rønning Rørteknikk",            "Bransje": "Rørlegger",                    "By": "Namsos",       "Manuelt problem": "Varelager og bestillinger uoversiktlig",                    "KABI-løsning": "Lager-Excel + bestillingsliste",       "Pris (NOK)": "5000-7000",   "Finn dem på": "Google Maps"},

    # FRILANSERE OG KONSULENTER (25)
    {"Nr": 31, "Bedriftsnavn": "Marte Vik Kommunikasjon",       "Bransje": "Kommunikasjonskonsulent",      "By": "Oslo",         "Manuelt problem": "Ukentlige statusrapporter lages manuelt",                   "KABI-løsning": "Rapport-generator",                    "Pris (NOK)": "4000-5000",   "Finn dem på": "LinkedIn"},
    {"Nr": 32, "Bedriftsnavn": "Anders Grønn Digital",          "Bransje": "Webutvikler",                  "By": "Bergen",       "Manuelt problem": "Tidregistrering og fakturering er separat",                 "KABI-løsning": "Excel timesheet + faktura",            "Pris (NOK)": "5000-6000",   "Finn dem på": "LinkedIn"},
    {"Nr": 33, "Bedriftsnavn": "Silje Holm PR",                 "Bransje": "PR-rådgiver",                  "By": "Oslo",         "Manuelt problem": "Medieklipp samles manuelt fra Google",                      "KABI-løsning": "Data-aggregator Excel",                "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 34, "Bedriftsnavn": "Kristoffer Dahl Foto",          "Bransje": "Fotograf",                     "By": "Stavanger",    "Manuelt problem": "Kontrakter sendes per e-post manuelt",                      "KABI-løsning": "Kontraktmal + email-automasjon",       "Pris (NOK)": "4000-5000",   "Finn dem på": "Instagram / LinkedIn"},
    {"Nr": 35, "Bedriftsnavn": "Tone Berg Regnskap",            "Bransje": "Regnskapskonsulent",           "By": "Trondheim",    "Manuelt problem": "Purringer på ubetalte fakturaer manuelt",                   "KABI-løsning": "Email-automasjon purringer",           "Pris (NOK)": "5000-6000",   "Finn dem på": "LinkedIn"},
    {"Nr": 36, "Bedriftsnavn": "Petter Sunde IT-løsninger",     "Bransje": "IT-konsulent",                 "By": "Oslo",         "Manuelt problem": "Ukerapporter til kunder tar 1-2 timer",                     "KABI-løsning": "Rapport-generator",                    "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 37, "Bedriftsnavn": "Maria Løkken Design",           "Bransje": "Grafisk designer",             "By": "Bergen",       "Manuelt problem": "Tilbudsskriving og fakturering i Word",                     "KABI-løsning": "Excel-fakturering + mal",              "Pris (NOK)": "4000-5000",   "Finn dem på": "LinkedIn / Instagram"},
    {"Nr": 38, "Bedriftsnavn": "Bjørn Aasen Advokat",           "Bransje": "Juridisk konsulent",           "By": "Oslo",         "Manuelt problem": "Dokumentsortering og arkivering manuelt",                  "KABI-løsning": "PDF-ekstraktor + Excel",               "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 39, "Bedriftsnavn": "Ingrid Nygård Coaching",        "Bransje": "Business coach",               "By": "Stavanger",    "Manuelt problem": "Bookinger og betaling håndteres per SMS",                  "KABI-løsning": "Booking-tracker Excel",                "Pris (NOK)": "4000-5000",   "Finn dem på": "Facebook / Instagram"},
    {"Nr": 40, "Bedriftsnavn": "Leif Eriksen Analyse",          "Bransje": "Dataanalytiker",               "By": "Trondheim",    "Manuelt problem": "Rapporter i Excel lages fra bunnen hver gang",              "KABI-løsning": "Rapport-mal + generator",              "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 41, "Bedriftsnavn": "Sofie Brun Innhold",            "Bransje": "Content creator",              "By": "Oslo",         "Manuelt problem": "Publiseringskalender finnes ikke",                          "KABI-løsning": "Innholds-planner Excel",               "Pris (NOK)": "4000-5000",   "Finn dem på": "LinkedIn / Instagram"},
    {"Nr": 42, "Bedriftsnavn": "Thomas Vik Revisjon",           "Bransje": "Revisor freelance",            "By": "Bergen",       "Manuelt problem": "Klientoversikt i hodet, ikke i system",                    "KABI-løsning": "CRM i Excel",                          "Pris (NOK)": "5000-7000",   "Finn dem på": "LinkedIn"},
    {"Nr": 43, "Bedriftsnavn": "Nora Haug Oversettelse",        "Bransje": "Oversetter",                   "By": "Oslo",         "Manuelt problem": "Leveringsfrister koordineres i e-post",                    "KABI-løsning": "Frist-tracker Excel",                  "Pris (NOK)": "4000-5000",   "Finn dem på": "LinkedIn"},
    {"Nr": 44, "Bedriftsnavn": "Ole Strand Strategi",           "Bransje": "Strategikonsulent",            "By": "Stavanger",    "Manuelt problem": "Møtenotater og oppfølgingspunkter mistes",                  "KABI-løsning": "Notat-to-Excel automatisering",        "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 45, "Bedriftsnavn": "Amalie Dahl Markedsføring",     "Bransje": "Markedsføringskonsulent",      "By": "Oslo",         "Manuelt problem": "Rapportering til kunder er tidkrevende",                   "KABI-løsning": "Rapport-generator",                    "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 46, "Bedriftsnavn": "Erik Thorsen Film",             "Bransje": "Videoprodusent",               "By": "Bergen",       "Manuelt problem": "Prosjektbudsjetter lages manuelt",                          "KABI-løsning": "Budsjett-Excel",                       "Pris (NOK)": "5000-6000",   "Finn dem på": "LinkedIn / Vimeo"},
    {"Nr": 47, "Bedriftsnavn": "Lise Bakken HR",                "Bransje": "HR-konsulent",                 "By": "Trondheim",    "Manuelt problem": "Onboarding-sjekkliste gjøres på nytt for hver klient",      "KABI-løsning": "Sjekkliste-generator",                 "Pris (NOK)": "5000-7000",   "Finn dem på": "LinkedIn"},
    {"Nr": 48, "Bedriftsnavn": "Jonas Moe Tekst",               "Bransje": "Tekstforfatter",               "By": "Oslo",         "Manuelt problem": "Ingen system for tilbud og oppfølging",                    "KABI-løsning": "Tilbuds-tracker + email-mal",          "Pris (NOK)": "4000-5000",   "Finn dem på": "LinkedIn"},
    {"Nr": 49, "Bedriftsnavn": "Hanne Olsen Event",             "Bransje": "Eventplanlegger",              "By": "Oslo",         "Manuelt problem": "Gjestlister og leverandøroversikt i Word",                 "KABI-løsning": "Excel-koordinering",                   "Pris (NOK)": "5000-6000",   "Finn dem på": "LinkedIn / Instagram"},
    {"Nr": 50, "Bedriftsnavn": "Fredrik Berg Finans",           "Bransje": "Finansrådgiver",               "By": "Stavanger",    "Manuelt problem": "Klientrapporter lages manuelt månedlig",                   "KABI-løsning": "Finansrapport-generator",              "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 51, "Bedriftsnavn": "Camilla Lie Utdanning",         "Bransje": "Kursholderinstruktør",         "By": "Bergen",       "Manuelt problem": "Deltakerlister og kursbevis manuelt",                      "KABI-løsning": "Excel + PDF-generator",                "Pris (NOK)": "5000-6000",   "Finn dem på": "Facebook / LinkedIn"},
    {"Nr": 52, "Bedriftsnavn": "Henrik Næss Sikkerhet",         "Bransje": "IT-sikkerhetskonsulent",       "By": "Oslo",         "Manuelt problem": "Sårbarhetsvurderinger dokumenteres manuelt",               "KABI-løsning": "Rapport-generator",                    "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 53, "Bedriftsnavn": "Vibeke Steen Terapi",           "Bransje": "Psykolog privat",              "By": "Stavanger",    "Manuelt problem": "Timeavtaler og notater ustrukturert",                       "KABI-løsning": "Booking-tracker Excel",                "Pris (NOK)": "4000-5000",   "Finn dem på": "Facebook / LinkedIn"},
    {"Nr": 54, "Bedriftsnavn": "Runar Elstad Rekruttering",     "Bransje": "Rekrutteringskonsulent",       "By": "Oslo",         "Manuelt problem": "CV-sortering og kandidatoversikt manuelt",                 "KABI-løsning": "Excel-sortering + tracker",            "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 55, "Bedriftsnavn": "Kine Gjerde Bok",               "Bransje": "Bokfører freelance",           "By": "Bergen",       "Manuelt problem": "Kvitteringer registreres manuelt fra bilder",              "KABI-løsning": "PDF-ekstraktor + Excel",               "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},

    # NETTBUTIKKER (25)
    {"Nr": 56, "Bedriftsnavn": "Norsk Kaffekultur AS",          "Bransje": "Kaffe og te nettbutikk",       "By": "Oslo",         "Manuelt problem": "Bestillinger kopieres manuelt fra e-post til Excel",        "KABI-løsning": "Bestillings-Excel + rapport",          "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook Ads / Instagram"},
    {"Nr": 57, "Bedriftsnavn": "Fjordlys Stearinlys",           "Bransje": "Håndlagde stearinlys",         "By": "Bergen",       "Manuelt problem": "Lagerhold oppdateres manuelt etter hvert salg",            "KABI-løsning": "Lager-tracker Excel",                  "Pris (NOK)": "5000-6000",   "Finn dem på": "Instagram / Etsy"},
    {"Nr": 58, "Bedriftsnavn": "Norsk Villmarksutstyr",         "Bransje": "Friluftslivsutstyr",           "By": "Trondheim",    "Manuelt problem": "Ingen oversikt over hvilke produkter tjener mest",          "KABI-løsning": "Lønnsomhets-analyse",                  "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook / Google Shopping"},
    {"Nr": 59, "Bedriftsnavn": "Bamses Hundemat",               "Bransje": "Hundematspesialisten",         "By": "Oslo",         "Manuelt problem": "Kundeservice-e-poster besvares manuelt én og én",          "KABI-løsning": "Email-automasjon FAQ",                 "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook / Instagram"},
    {"Nr": 60, "Bedriftsnavn": "Solsikke Barneklær",            "Bransje": "Barneklær og tilbehør",        "By": "Stavanger",    "Manuelt problem": "Returer registreres på papir",                              "KABI-løsning": "Retur-tracker Excel",                  "Pris (NOK)": "4000-5000",   "Finn dem på": "Instagram / Facebook"},
    {"Nr": 61, "Bedriftsnavn": "Norgesgave Nettbutikk",         "Bransje": "Norske gaveprodukter",         "By": "Bergen",       "Manuelt problem": "Pakkesedler skrives ut manuelt",                            "KABI-løsning": "Pakkseddel-generator",                 "Pris (NOK)": "5000-6000",   "Finn dem på": "Google Shopping"},
    {"Nr": 62, "Bedriftsnavn": "Trehjørnet Møbler",             "Bransje": "Håndlagde møbler",             "By": "Oslo",         "Manuelt problem": "Bestillingsbekreftelser sendes manuelt",                    "KABI-løsning": "Email-automasjon",                     "Pris (NOK)": "5000-7000",   "Finn dem på": "Instagram / Pinterest"},
    {"Nr": 63, "Bedriftsnavn": "Kystens Smykker",               "Bransje": "Håndlagde smykker",            "By": "Ålesund",      "Manuelt problem": "Ingen månedlig salgsrapport",                               "KABI-løsning": "Salgsrapport-generator",               "Pris (NOK)": "4000-5000",   "Finn dem på": "Etsy / Instagram"},
    {"Nr": 64, "Bedriftsnavn": "Villbær Kosmetikk",             "Bransje": "Naturlig hudpleie",            "By": "Tromsø",       "Manuelt problem": "Råvarelager ikke sporet digitalt",                          "KABI-løsning": "Lager-tracker Excel",                  "Pris (NOK)": "5000-6000",   "Finn dem på": "Instagram / Facebook"},
    {"Nr": 65, "Bedriftsnavn": "Havremel Bakeri",               "Bransje": "Norske bakervarer online",     "By": "Fredrikstad",  "Manuelt problem": "Ukentlige bestillinger koordineres via SMS",                "KABI-løsning": "Bestillings-Excel + email",            "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook"},
    {"Nr": 66, "Bedriftsnavn": "Norgespil Brettspill",          "Bransje": "Norskproduserte brettspill",   "By": "Drammen",      "Manuelt problem": "Kundeanmeldelser samles ikke systematisk",                  "KABI-løsning": "Anmeldelses-tracker Excel",            "Pris (NOK)": "4000-5000",   "Finn dem på": "Facebook / BGG"},
    {"Nr": 67, "Bedriftsnavn": "Tindetoppen Sport",             "Bransje": "Ski og friluftsliv",           "By": "Lillehammer",  "Manuelt problem": "Sesongvarer ikke sporet",                                   "KABI-løsning": "Sesong-lager Excel",                   "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook / Google Shopping"},
    {"Nr": 68, "Bedriftsnavn": "Slektsgård Honning",            "Bransje": "Lokal honning og syltetøy",    "By": "Hamar",        "Manuelt problem": "Faktura til grossister lages manuelt",                      "KABI-løsning": "Excel-fakturering",                    "Pris (NOK)": "4000-5000",   "Finn dem på": "Facebook / Finn.no"},
    {"Nr": 69, "Bedriftsnavn": "Urban Plantebutikk",            "Bransje": "Innendørsplanter nett",        "By": "Oslo",         "Manuelt problem": "Ingen automatisk påminnelse for vanningsabonnement",        "KABI-løsning": "Email-automasjon",                     "Pris (NOK)": "6000-8000",   "Finn dem på": "Instagram"},
    {"Nr": 70, "Bedriftsnavn": "Fiskers Krok",                  "Bransje": "Sportsfiskeutstyr",            "By": "Bodø",         "Manuelt problem": "Leverandørfakturaer registreres manuelt",                   "KABI-løsning": "PDF-ekstraktor + Excel",               "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook"},
    {"Nr": 71, "Bedriftsnavn": "Norsk Ull & Garn",              "Bransje": "Strikketilbehør",              "By": "Stavanger",    "Manuelt problem": "Ingen segmentering av kundeliste",                          "KABI-løsning": "Kundesegmentering Excel",              "Pris (NOK)": "5000-7000",   "Finn dem på": "Ravelry / Instagram"},
    {"Nr": 72, "Bedriftsnavn": "Gokart Gear",                   "Bransje": "Racing og motorsport tilbehør","By": "Porsgrunn",    "Manuelt problem": "Bestillingsstatistikk finnes ikke",                         "KABI-løsning": "Statistikk-rapport Excel",             "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook"},
    {"Nr": 73, "Bedriftsnavn": "Regnbue Barneleker",            "Bransje": "Pedagogisk leketøy",           "By": "Bergen",       "Manuelt problem": "Lagerstatus oppdateres for sent",                           "KABI-løsning": "Lager-alarm Excel",                    "Pris (NOK)": "5000-6000",   "Finn dem på": "Instagram / Facebook"},
    {"Nr": 74, "Bedriftsnavn": "Fjellvind Klær",                "Bransje": "Utendørsklær merke",           "By": "Trondheim",    "Manuelt problem": "Sesongkolleksjon-planlegging i hodet",                      "KABI-løsning": "Planleggings-Excel",                   "Pris (NOK)": "6000-8000",   "Finn dem på": "Instagram"},
    {"Nr": 75, "Bedriftsnavn": "Hav og Hei Kunst",              "Bransje": "Norske kunsttrykk",            "By": "Oslo",         "Manuelt problem": "Fraktoppfølging gjøres manuelt",                            "KABI-løsning": "Frakt-tracker Excel",                  "Pris (NOK)": "4000-5000",   "Finn dem på": "Instagram / Etsy"},
    {"Nr": 76, "Bedriftsnavn": "Naturens Apotek",               "Bransje": "Naturmedisin nettbutikk",      "By": "Stavanger",    "Manuelt problem": "Abonnementskunder ikke sporet",                             "KABI-løsning": "Abonnements-tracker",                  "Pris (NOK)": "6000-8000",   "Finn dem på": "Facebook / Instagram"},
    {"Nr": 77, "Bedriftsnavn": "Norsk Tekno Gadgets",           "Bransje": "Teknologiprodukter",           "By": "Oslo",         "Manuelt problem": "Prisoppdateringer gjøres manuelt",                          "KABI-løsning": "Prisoppdaterings-verktøy",             "Pris (NOK)": "8000-10000",  "Finn dem på": "Google Shopping"},
    {"Nr": 78, "Bedriftsnavn": "Sjømat Direkte",                "Bransje": "Fisk og skalldyr online",      "By": "Bergen",       "Manuelt problem": "Bestillinger og frakt koordineres manuelt",                 "KABI-løsning": "Bestillings-Excel + email",            "Pris (NOK)": "8000-10000",  "Finn dem på": "Facebook"},
    {"Nr": 79, "Bedriftsnavn": "Vinterbær Syltetøy",            "Bransje": "Norsk syltetøy og sauser",     "By": "Molde",        "Manuelt problem": "Ingen utsendelse av nyhetsbrev automatisk",                 "KABI-løsning": "Email-automasjon liste",               "Pris (NOK)": "5000-6000",   "Finn dem på": "Facebook / Instagram"},
    {"Nr": 80, "Bedriftsnavn": "Lykkepose Abonnement",          "Bransje": "Abonnementsgaver",             "By": "Oslo",         "Manuelt problem": "Abonnenter registreres i regneark manuelt",                "KABI-løsning": "Abonnements-tracker Excel",            "Pris (NOK)": "6000-8000",   "Finn dem på": "Instagram / Facebook"},

    # REGNSKAPSFØRERE OG ADVOKATER (20)
    {"Nr": 81,  "Bedriftsnavn": "Regnskap Vest AS",              "Bransje": "Regnskapskontor",              "By": "Bergen",        "Manuelt problem": "Månedlig rapportpakke til klienter tar lang tid",          "KABI-løsning": "Finansrapport-generator",              "Pris (NOK)": "10000-15000", "Finn dem på": "LinkedIn"},
    {"Nr": 82,  "Bedriftsnavn": "Advokat Persen & Co",           "Bransje": "Advokatfirma",                 "By": "Oslo",          "Manuelt problem": "Dokumenter fra klienter sorteres manuelt",                 "KABI-løsning": "PDF-ekstraktor + Excel",               "Pris (NOK)": "10000-15000", "Finn dem på": "LinkedIn"},
    {"Nr": 83,  "Bedriftsnavn": "Aker Revisjon AS",              "Bransje": "Revisjonsselskap",             "By": "Stavanger",     "Manuelt problem": "Årsoppgjørsdata samles fra mange Excel-filer",             "KABI-løsning": "Excel-aggregering",                    "Pris (NOK)": "10000-15000", "Finn dem på": "LinkedIn"},
    {"Nr": 84,  "Bedriftsnavn": "Bergstad Regnskap",             "Bransje": "Regnskapskontor",              "By": "Trondheim",     "Manuelt problem": "Purrebrev skrives og sendes manuelt",                      "KABI-løsning": "Email-automasjon",                     "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn / Google Maps"},
    {"Nr": 85,  "Bedriftsnavn": "Solvik Advokater",              "Bransje": "Advokatfirma",                 "By": "Oslo",          "Manuelt problem": "Timeregistrering og fakturering ikke koblet",              "KABI-løsning": "Excel timesheet + faktura",            "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 86,  "Bedriftsnavn": "Tallmaster Regnskap",           "Bransje": "Enkeltpersonsregnskapsbyå",   "By": "Bergen",        "Manuelt problem": "Klientliste i hodene til de ansatte",                      "KABI-løsning": "CRM i Excel",                          "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 87,  "Bedriftsnavn": "Nord-Norsk Revisjon",           "Bransje": "Revisjonsselskap",             "By": "Tromsø",        "Manuelt problem": "Årsberetninger skrives fra bunnen hver gang",              "KABI-løsning": "Rapport-mal + generator",              "Pris (NOK)": "10000-15000", "Finn dem på": "LinkedIn"},
    {"Nr": 88,  "Bedriftsnavn": "Lovdata Advokat AS",            "Bransje": "Advokatfirma",                 "By": "Trondheim",     "Manuelt problem": "Saksarkiv er ustrukturert",                                "KABI-løsning": "Saks-tracker Excel + PDF",             "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 89,  "Bedriftsnavn": "Vestland Regnskap",             "Bransje": "Regnskapskontor",              "By": "Stavanger",     "Manuelt problem": "Klientrapporter per e-post, ingen mal",                    "KABI-løsning": "Rapport-mal + generator",              "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 90,  "Bedriftsnavn": "Kristiansen Jus",               "Bransje": "Advokat enkeltperson",         "By": "Oslo",          "Manuelt problem": "Kontraktutkast startes fra bunnen",                        "KABI-løsning": "Kontraktmal-generator",                "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 91,  "Bedriftsnavn": "Fjord Regnskap & Skatt",        "Bransje": "Regnskapskontor",              "By": "Ålesund",       "Manuelt problem": "MVA-rapportering manuelt kompilert",                       "KABI-løsning": "MVA-rapport Excel",                    "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 92,  "Bedriftsnavn": "Helgeland Revisjon",            "Bransje": "Revisjonsselskap",             "By": "Bodø",          "Manuelt problem": "Intern rapportering til styre manuelt",                    "KABI-løsning": "Styresrapport-generator",              "Pris (NOK)": "10000-15000", "Finn dem på": "LinkedIn"},
    {"Nr": 93,  "Bedriftsnavn": "Bymann Advokat",                "Bransje": "Advokatfirma",                 "By": "Drammen",       "Manuelt problem": "Klientfakturaer lages i Word",                             "KABI-løsning": "Faktura-generator Excel",              "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 94,  "Bedriftsnavn": "Østnorsk Regnskap AS",          "Bransje": "Regnskapskontor",              "By": "Fredrikstad",   "Manuelt problem": "Månedlig lønnsoversikt til klienter manuell",              "KABI-løsning": "Lønnsrapport-generator",               "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 95,  "Bedriftsnavn": "Juris Finans Oslo",             "Bransje": "Regnskaps- og juridisk rådgivning", "By": "Oslo",    "Manuelt problem": "Klientoppfølging etter prosjekter mangler",                "KABI-løsning": "Email-automasjon + tracker",           "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 96,  "Bedriftsnavn": "Hordaland Regnskapsservice",    "Bransje": "Regnskapskontor",              "By": "Bergen",        "Manuelt problem": "Regnskapsdata fra flere systemer samles manuelt",          "KABI-løsning": "Excel-aggregering",                    "Pris (NOK)": "10000-15000", "Finn dem på": "LinkedIn"},
    {"Nr": 97,  "Bedriftsnavn": "Lunde Advokatkontor",           "Bransje": "Advokatfirma",                 "By": "Stavanger",     "Manuelt problem": "Saksnotatene er ufullstendige og rotete",                  "KABI-løsning": "Notat-mal + PDF-ekstraktor",           "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 98,  "Bedriftsnavn": "Finans & Lov AS",               "Bransje": "Kombinert regnskap og jus",    "By": "Oslo",          "Manuelt problem": "Ingen strukturert onboarding av nye klienter",             "KABI-løsning": "Onboarding-mal + tracker",             "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
    {"Nr": 99,  "Bedriftsnavn": "Fjellstad Regnskap",            "Bransje": "Enkeltpersonsregnskapsbyrå",   "By": "Hamar",         "Manuelt problem": "Ingen automatisk varsling om frister",                     "KABI-løsning": "Frist-tracker + email-alarm",          "Pris (NOK)": "6000-8000",   "Finn dem på": "LinkedIn"},
    {"Nr": 100, "Bedriftsnavn": "Solberg & Partnere Advokater",  "Bransje": "Advokatfirma",                 "By": "Kristiansand",  "Manuelt problem": "Timerapporter til klienter lages manuelt",                 "KABI-løsning": "Timerapport-generator",                "Pris (NOK)": "8000-10000",  "Finn dem på": "LinkedIn"},
]

print(f"Laster {len(kunder)} kunder...")

# ── 2. EXCEL ──────────────────────────────────────────────────────────────────

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(OUTPUT_DIR, "kundeliste.xlsx")
MD_PATH    = os.path.join(OUTPUT_DIR, "outreach-liste.md")

print("Genererer Excel-fil...")

df = pd.DataFrame(kunder)
df.to_excel(EXCEL_PATH, index=False, engine="openpyxl")

wb = load_workbook(EXCEL_PATH)
ws = wb.active
ws.title = "Kundeliste"

# Styles
header_fill  = PatternFill("solid", fgColor="1F4E79")
header_font  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
row_fill_alt = PatternFill("solid", fgColor="DEE9F1")
row_fill_wht = PatternFill("solid", fgColor="FFFFFF")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

col_widths = {
    "Nr":             5,
    "Bedriftsnavn":   30,
    "Bransje":        20,
    "By":             18,
    "Manuelt problem":40,
    "KABI-løsning":   30,
    "Pris (NOK)":     18,
    "Finn dem på":    25,
}

# Apply header formatting
for col_idx, col_name in enumerate(df.columns, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill   = header_fill
    cell.font   = header_font
    cell.alignment = center_align
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

# Apply row formatting
for row_idx in range(2, len(kunder) + 2):
    fill = row_fill_alt if (row_idx % 2 == 0) else row_fill_wht
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.alignment = center_align if col_idx == 1 else left_align

# Freeze top row
ws.freeze_panes = "A2"

# Row height
ws.row_dimensions[1].height = 22
for row_idx in range(2, len(kunder) + 2):
    ws.row_dimensions[row_idx].height = 18

wb.save(EXCEL_PATH)
print(f"  Lagret: {EXCEL_PATH}")

# ── 3. MARKDOWN OUTREACH ──────────────────────────────────────────────────────

print("Genererer outreach-meldinger...")

# Personalized Norwegian outreach messages for all 100 customers
outreach_messages = {
    1:   "Hei Hansen Rør & Varme! Vi ser at dere bruker mye tid på manuell fakturaregistrering i Excel. KABI Automation kan automatisere hele prosessen og generere ferdige rapporter på sekunder – slik at dere kan fokusere på rørleggerarbeidet, ikke regneark.",
    2:   "Hei Nordvik Elektro! Tidsskjema og timelister på papir er en klassisk tidtyv for elektriker-firmaer. Vi kan sette opp et digitalt Excel-system som registrerer timer automatisk og gir dere oversikt over alle jobber uten ekstra manuelt arbeid.",
    3:   "Hei Bakke Snekker & Bygg! Å bruke 2-3 timer per kunde på tilbudsskriving er dyrt. KABI Automation lager en e-postmal og automatisk tilbudsgenerator som kutter den tiden til minutter – så dere kan ta flere oppdrag.",
    4:   "Hei Solberg VVS! Manuelle purringer på ubetalte fakturaer er tidkrevende og flaut. Vi setter opp automatiske e-postpåminnelser som sender seg selv – dere får betalt raskere uten å løfte en finger.",
    5:   "Hei Østlund Maler! Rapportskriving etter hvert oppdrag tar unødvendig tid. KABI Automation bygger en rapport-generator som fyller ut standardinfo automatisk – dere trenger bare å godkjenne og sende.",
    6:   "Hei Fjord Elektrikere! Manuelle materialbestillinger per telefon og e-post skaper rot og forsinkelser. Vi lager et Excel-system som samler bestillingsbehovene og sender dem strukturert til leverandørene.",
    7:   "Hei Strand Tømrer! Å holde styr på materiallager i hodet er risikabelt – det gir feil og ekstra turer til butikken. KABI lager en enkel lager-tracker i Excel som alltid viser hva dere har og hva som må bestilles.",
    8:   "Hei Viken Rørservice! Uten systematisk kundeoppfølging mister dere gjentakende oppdrag. Vi setter opp et enkelt CRM i Excel med e-postmaler som holder kundene varme – automatisk.",
    9:   "Hei Haugen Byggservice! Sen fakturering er en av de vanligste grunnene til likviditetsproblemer i håndverksbransjen. KABI lager en automatisk faktura-påminner som varsler dere når det er tid for å fakturere.",
    10:  "Hei Lunde Elektro! HMS-rapporter som skrives manuelt er både tidkrevende og feilutsatt. Vi bygger en rapport-generator som produserer ferdig formaterte HMS-dokumenter på under ett minutt.",
    11:  "Hei Aas & Sønner Rør! Uten oversikt over aktive jobber er det lett å gå glipp av frister eller dobbeltbooke ressurser. KABI lager en jobb-tracker i Excel som gir dere full kontroll over alle pågående oppdrag.",
    12:  "Hei Berg Maler og Fasade! Kundekommunikasjon via SMS og lapper er uprofesjonelt og lett å miste. Vi setter opp et e-postsystem med maler som gjør all kommunikasjon sporbar og profesjonell.",
    13:  "Hei Dahl VVS og Baderom! Feil i timepriskalkylasjon kan koste dere tusenvis per prosjekt. KABI lager et Excel-verktøy som automatisk kalkulerer korrekt timepris basert på kostnad og ønsket margin.",
    14:  "Hei Nilsen Elektriske! Uten servicehistorikk på kundene mister dere verdifull informasjon ved hvert nytt besøk. Vi bygger en enkel historikk-database i Excel så dere alltid vet hva som er gjort og når.",
    15:  "Hei Kvam Bygg & Renovering! Telefonkoordinering av alle avtaler er tidkrevende og skaper lett misforståelser. KABI lager et koordinerings-Excel med maler som gir alle involverte klar oversikt automatisk.",
    16:  "Hei Rørvik Sanitær! Faktura i Word uten system er en oppskrift på rot og sen betaling. Vi setter opp en strukturert Excel-fakturamodell som er rask, profesjonell og enkel å bruke.",
    17:  "Hei Lie Elektro og Sikkerhet! Manuelle sjekklister for hvert oppdrag tar tid og kan inneholde feil. KABI lager en sjekkliste-generator som produserer oppdragsspesifikke lister automatisk – klare til bruk.",
    18:  "Hei Munthe Tømrer! Timeregistrering på papir er upålitelig og gjør fakturering vanskeligere. Vi bygger et Excel timesheet med automatisk rapport-funksjon som gjør lønnsutregning og fakturering enkelt.",
    19:  "Hei Holm Maling og Sparkling! Uten lønnsomhetsoversikt vet dere ikke hvilke kunder og oppdragstyper som faktisk betaler seg. KABI lager en analyse-modell i Excel som viser nøyaktig hva som lønner seg.",
    20:  "Hei Elgaard Rørlegging! Sene tilbud betyr tapte oppdrag – kunden velger den som svarer raskest. Vi lager en e-postmal for hurtigtilbud som gjør at dere kan sende profesjonelle tilbud på minutter.",
    21:  "Hei Thorsen Elektro! Manuell lønnsberegning til ansatte er tidkrevende og feilutsatt. KABI lager en Excel-kalkulator som automatisk regner ut lønn basert på timer, satser og tillegg.",
    22:  "Hei Berge Byggtjenester! Uten et system for å følge opp leads mister dere potensielle oppdrag. Vi setter opp en lead-tracker i Excel som minner dere om hvem som skal kontaktes og når.",
    23:  "Hei Engen VVS Service! Glemt garantioppfølging kan bli kostbart både økonomisk og omdømmemessig. KABI lager en garanti-tracker som automatisk varsler dere før garantifrister utløper.",
    24:  "Hei Myhre Elektrikeren! Ukjent timeforbruk per prosjekt gjør det umulig å prise fremtidige jobber riktig. Vi bygger en prosjekt-timer Excel som gir dere nøyaktig oversikt over timebruk per oppdrag.",
    25:  "Hei Solli Snekker & Interiør! Uten standardiserte tilbudsmaler bruker dere unødvendig tid på hvert tilbud. KABI lager en tilbuds-generator som produserer profesjonelle, skreddersydde tilbud på minutter.",
    26:  "Hei Birkeland Rør! Manuelle SMS-varsler til kunder tar tid og glemmes lett. Vi automatiserer kundekommunikasjonen med e-postmaler som sender seg selv ved bestemte hendelser.",
    27:  "Hei Finstad Maler! Uten digital jobbregistrering er det vanskelig å dokumentere hva som er gjort og fakturere korrekt. KABI lager en enkel jobb-logg i Excel som er rask å fylle ut og enkel å bruke.",
    28:  "Hei Løken Elektro! Manuelle forsikringsrapporter tar tid og øker risikoen for feil. Vi bygger en rapport-generator som produserer korrekt formaterte dokumenter automatisk – klare for innsending.",
    29:  "Hei Wold Bygg & Rehabilitering! Prosjektbudsjett som sprekker uten varsel gir ubehagelige overraskelser. KABI lager en budsjett-tracker i Excel som varsler dere automatisk når dere nærmer deg grensen.",
    30:  "Hei Rønning Rørteknikk! Uoversiktlig varelager og bestillinger fører til forsinkelser og ekstra kostnader. Vi lager et strukturert lager-Excel med automatisk bestillingsliste som holder dere à jour.",
    31:  "Hei Marte Vik! Ukentlige statusrapporter som lages manuelt spiser opp verdifull fakturerbar tid. KABI Automation bygger en rapport-generator som setter sammen ferdig formaterte oppdateringer automatisk.",
    32:  "Hei Anders Grønn Digital! Separat tidregistrering og fakturering er tidkrevende og øker risikoen for feil. Vi kobler timesheet og fakturamodell i ett Excel-system – logg timer, generer faktura med ett klikk.",
    33:  "Hei Silje Holm PR! Manuell innsamling av medieklipp fra Google er en enorm tidtyv. KABI lager en data-aggregator i Excel som samler og sorterer medieomtale automatisk, klar for rapportering til klienter.",
    34:  "Hei Kristoffer Dahl Foto! Manuell kontraktsutsending per e-post er uprofesjonelt og lett å glemme. Vi lager standardiserte kontraktmaler og automatiserer utsendingen, slik at du alltid er dekket.",
    35:  "Hei Tone Berg Regnskap! Manuelle purringer på ubetalte fakturaer tar tid du heller kan bruke på klienter. KABI setter opp automatiske e-postpåminnelser i riktig tone – profesjonelt og uten ekstra arbeid.",
    36:  "Hei Petter Sunde IT! Ukerapporter som tar 1-2 timer er 1-2 timer du ikke fakturerer for. Vi bygger en rapport-generator som samler data automatisk og produserer klientrapporter på minutter.",
    37:  "Hei Maria Løkken Design! Tilbud og fakturaer i Word uten system gir rot og uprofesjonelt uttrykk. KABI lager en Excel-basert fakturamodell med tilbudsmaler som gjør hele prosessen rask og konsistent.",
    38:  "Hei Bjørn Aasen Advokat! Manuell dokumentsortering og arkivering er en risiko i juridisk arbeid. Vi bygger en PDF-ekstraktor med Excel-indeks som strukturerer alle dokumenter automatisk og gjør dem søkbare.",
    39:  "Hei Ingrid Nygård Coaching! Bookinger og betaling via SMS er uprofesjonelt og vanskelig å holde oversikt over. KABI lager en strukturert booking-tracker i Excel som gir deg full oversikt uten ekstra chaos.",
    40:  "Hei Leif Eriksen Analyse! Å starte Excel-rapporter fra bunnen hver gang er ineffektivt for en dataanalytiker. Vi bygger gjenbrukbare rapport-maler med automatisk datafylling som kutter produksjonstiden drastisk.",
    41:  "Hei Sofie Brun Innhold! Uten en publiseringskalender er det lett å miste kontroll over hva som skal ut når. KABI lager en innholds-planner i Excel som gir deg full oversikt over alle kanaler og datoer.",
    42:  "Hei Thomas Vik Revisjon! Klientoversikt i hodet er risikabelt – ett glemt møte kan skade forholdet. Vi setter opp et enkelt CRM i Excel med påminnelser som holder alle klienter organisert.",
    43:  "Hei Nora Haug Oversettelse! Leveringsfrister koordinert i e-post er lett å miste og stressende å følge opp. KABI lager en frist-tracker i Excel som gir deg klar oversikt over alle pågående prosjekter.",
    44:  "Hei Ole Strand Strategi! Møtenotater og oppfølgingspunkter som mistes koster deg troverdighet hos klienter. Vi automatiserer notat-til-Excel-prosessen så alle aksjoner fanges opp og følges opp systematisk.",
    45:  "Hei Amalie Dahl Markedsføring! Tidkrevende klientrapportering er penger du ikke tjener. KABI bygger en rapport-generator som henter data automatisk og produserer profesjonelle markedsrapporter på minutter.",
    46:  "Hei Erik Thorsen Film! Manuelle prosjektbudsjetter er upresise og tar tid. Vi lager et Excel-budsjett skreddersydd for videoproduksjon – med automatisk beregning av kostnader, marginer og klientfaktura.",
    47:  "Hei Lise Bakken HR! Å gjøre onboarding-sjekklisten på nytt for hver klient er unødvendig dobbeltarbeid. KABI lager en sjekkliste-generator som automatisk tilpasser onboarding-løpet til hver ny klient.",
    48:  "Hei Jonas Moe Tekst! Uten et system for tilbud og oppfølging mister du potensielle oppdrag. Vi lager en tilbuds-tracker med e-postmaler som holder deg oppdatert på alle leads og sikrer god oppfølging.",
    49:  "Hei Hanne Olsen Event! Gjestlister og leverandøroversikt i Word er kaotisk og feilutsatt. KABI lager et Excel-koordineringsverktøy som samler alt på ett sted – gjester, leverandører, frister og status.",
    50:  "Hei Fredrik Berg Finans! Manuelle månedlige klientrapporter er dyrt. Vi bygger en finansrapport-generator som henter og strukturerer data automatisk – slik at rapportene er klare på minutter, ikke timer.",
    51:  "Hei Camilla Lie Utdanning! Manuelle deltakerlister og kursbevis tar tid du burde bruke på å undervise. KABI lager et Excel-system som automatisk genererer kursbevis og deltakerlister basert på påmeldingsdata.",
    52:  "Hei Henrik Næss Sikkerhet! Manuell dokumentasjon av sårbarhetsvurderinger er tidkrevende og risikabelt. Vi bygger en rapport-generator som produserer strukturerte, profesjonelle sikkerhetsdokumenter automatisk.",
    53:  "Hei Vibeke Steen Terapi! Ustrukturerte timeavtaler og notater skaper stress og kan gå ut over klientarbeidet. KABI lager en enkel booking-tracker i Excel som gir full oversikt uten tunge systemer.",
    54:  "Hei Runar Elstad Rekruttering! Manuell CV-sortering og kandidatoversikt er en enorm tidtyv. Vi bygger et Excel-sorteringssystem som automatisk kategoriserer kandidater og gir deg oversikt over hele pipelinen.",
    55:  "Hei Kine Gjerde Bok! Manuell registrering av kvitteringer fra bilder er tidkrevende og feilutsatt. KABI lager en PDF-ekstraktor som automatisk leser av og registrerer kvitteringsdata i Excel.",
    56:  "Hei Norsk Kaffekultur! Manuell kopiering av bestillinger fra e-post til Excel er ineffektivt og feilutsatt. KABI lager et bestillings-Excel som automatisk henter og strukturerer innkommende ordrer med rapport.",
    57:  "Hei Fjordlys Stearinlys! Lagerhold som oppdateres manuelt etter hvert salg er tidkrevende og upålitelig. Vi lager en lager-tracker i Excel som alltid er oppdatert og varsler dere når varer er i ferd med å gå tom.",
    58:  "Hei Norsk Villmarksutstyr! Uten oversikt over hvilke produkter som tjener mest er det vanskelig å prioritere riktig. KABI lager en lønnsomhetsanalyse-modell som viser nøyaktig hvilke produkter dere bør satse på.",
    59:  "Hei Bamses Hundemat! Manuelle svar på kundeservice-e-poster én og én er en enorm tidtyv. Vi lager automatiserte e-postsvar for de vanligste spørsmålene – kundene får rask hjelp, dere sparer tid.",
    60:  "Hei Solsikke Barneklær! Returer på papir gir dårlig oversikt og vanskeliggjør statistikk. KABI lager en retur-tracker i Excel som registrerer alle returer digitalt og gir dere god kontroll og rapportering.",
    61:  "Hei Norgesgave Nettbutikk! Manuelle pakkesedler er tidkrevende og feilutsatt ved høy ordre-volum. Vi lager en pakkseddel-generator som automatisk produserer ferdig utfylte pakkesedler fra ordredata.",
    62:  "Hei Trehjørnet Møbler! Manuelle bestillingsbekreftelser tar tid og øker risikoen for at kunder ikke får svar. KABI automatiserer utsendingen av profesjonelle bekreftelsese-poster så kunder alltid er informert.",
    63:  "Hei Kystens Smykker! Uten månedlig salgsrapport er det vanskelig å forstå hvilke produkter og kanaler som fungerer. Vi lager en salgsrapport-generator som automatisk sammenstiller salgsdata fra alle kanaler.",
    64:  "Hei Villbær Kosmetikk! Råvarelager som ikke spores digitalt gir produksjonsstopp og unødvendige hastebestillinger. KABI lager en lager-tracker i Excel som gir deg full kontroll over alle ingredienser og materialer.",
    65:  "Hei Havremel Bakeri! Ukentlige bestillinger via SMS er uprofesjonelt og vanskelig å holde oversikt over. Vi lager et bestillings-Excel med automatisk e-postbekreftelse som gjør hele prosessen strukturert og enkel.",
    66:  "Hei Norgespil Brettspill! Kundeanmeldelser som ikke samles systematisk er en gullgruve som ikke utnyttes. KABI lager en anmeldelses-tracker i Excel som samler og kategoriserer feedback fra alle plattformer.",
    67:  "Hei Tindetoppen Sport! Sesongvarer som ikke spores gir enten overfylte lagre eller utsolgte bestselgere. Vi lager et sesong-lager Excel som hjelper dere planlegge innkjøp og unngå kostbare feiltrinn.",
    68:  "Hei Slektsgård Honning! Manuelle fakturaer til grossister er tidkrevende og risikabelt. KABI lager et Excel-faktureringssystem skreddersydd for grossist-salg – raskt, profesjonelt og uten feil.",
    69:  "Hei Urban Plantebutikk! Uten automatiske påminnelser til abonnenter risikerer dere tap av kunder. Vi setter opp automatiserte e-postpåminnelser om vanningsabonnementet som holder kundene engasjert.",
    70:  "Hei Fiskers Krok! Manuell registrering av leverandørfakturaer fra PDF er tidkrevende og feilutsatt. KABI lager en PDF-ekstraktor som automatisk leser og overfører fakturainfo til Excel-regnskapet.",
    71:  "Hei Norsk Ull & Garn! Uten segmentering av kundelisten sender dere samme budskap til alle og mister relevans. Vi lager et kundesegmenteringsverktøy i Excel som deler listen inn etter kjøpsatferd og interesser.",
    72:  "Hei Gokart Gear! Uten bestillingsstatistikk er det vanskelig å ta gode innkjøps- og markedsføringsbeslutninger. KABI lager en statistikk-rapport Excel som gir dere klare innsikter om salg, trender og kunder.",
    73:  "Hei Regnbue Barneleker! Sen lagerstatus-oppdatering fører til at dere selger varer som er utsolgt. Vi lager en lager-alarm Excel som varsler automatisk når beholdningen faller under et definert minstenivå.",
    74:  "Hei Fjellvind Klær! Sesongkolleksjon-planlegging i hodet er risikabelt og gjør det vanskelig å samarbeide. KABI lager et planleggings-Excel som strukturerer hele kolleksjonsprosessen fra idé til lansering.",
    75:  "Hei Hav og Hei Kunst! Manuell fraktoppfølging er tidkrevende og frustrerende for kunder som vil ha oppdateringer. Vi lager en frakt-tracker i Excel som gir full oversikt over alle pakker og leveringsstatus.",
    76:  "Hei Naturens Apotek! Abonnementskunder som ikke spores gir dårlig oversikt og økt frafall. KABI lager en abonnements-tracker som automatisk viser hvem som er aktive, hvem som skal fornye og hvem som er i fare for å churne.",
    77:  "Hei Norsk Tekno Gadgets! Manuelle prisoppdateringer er tidkrevende og øker risikoen for feil på Google Shopping. Vi lager et prisoppdateringsverktøy som gjør masseoppdateringer raskt og strukturert.",
    78:  "Hei Sjømat Direkte! Manuell koordinering av bestillinger og frakt for fersk sjømat gir for stor risiko for feil. KABI lager et bestillings-Excel med automatisk e-postbekreftelse og fraktkoordinering.",
    79:  "Hei Vinterbær Syltetøy! Uten automatisk nyhetsbrev mister dere en viktig kanal for gjentakende salg. Vi setter opp en e-postautomasjon med ny kundliste som sender nyhetsbrev uten manuelt arbeid.",
    80:  "Hei Lykkepose Abonnement! Manuell abonnentregistrering i regneark er uholdbart i lengden. KABI lager en automatisert abonnements-tracker som holder styr på alle kunder, fornyelser og betalingsstatus.",
    81:  "Hei Regnskap Vest! Månedlig rapportpakke til klienter som tar lang tid er direkte lønnsomt å automatisere. KABI bygger en finansrapport-generator skreddersydd for regnskapskontor – produser alle klientrapporter på minutter.",
    82:  "Hei Advokat Persen & Co! Manuell dokumentsortering er risikabelt i juridisk arbeid og stjeler fakturerbar tid. Vi lager en PDF-ekstraktor med Excel-indeks som automatisk klassifiserer og arkiverer innkommende dokumenter.",
    83:  "Hei Aker Revisjon! Å samle årsoppgjørsdata fra mange Excel-filer manuelt er tidkrevende og feilutsatt. KABI lager et aggregeringsverktøy som automatisk konsoliderer data fra alle klientfiler til en strukturert rapport.",
    84:  "Hei Bergstad Regnskap! Manuelle purrebrev tar tid og gir inkonsistent kommunikasjon til klienter. Vi setter opp automatiserte purringer med riktig tone og timing – profesjonelt og uten ekstra arbeid.",
    85:  "Hei Solvik Advokater! Timeregistrering og fakturering som ikke er koblet gir tapt inntekt og ekstraarbeid. KABI lager et integrert Excel-system der timer automatisk kobles til fakturaen – ingen inntekt mistes.",
    86:  "Hei Tallmaster Regnskap! Klientliste som bare eksisterer i de ansattes hoder er en forretningsrisiko. Vi bygger et CRM i Excel som strukturerer all klientinfo, historikk og oppfølgingspunkter på ett sted.",
    87:  "Hei Nord-Norsk Revisjon! Årsberetninger som skrives fra bunnen hver gang er unødvendig dobbeltarbeid. KABI lager gjenbrukbare rapport-maler med automatisk datafylling som kutter produksjonstiden drastisk.",
    88:  "Hei Lovdata Advokat! Ustrukturert saksarkiv skaper forsinkelser og øker risikoen for å miste viktig informasjon. Vi lager en saks-tracker Excel med PDF-indeksering som gir full oversikt og rask gjenfinning.",
    89:  "Hei Vestland Regnskap! Klientrapporter uten mal gir inkonsistent kvalitet og tar unødvendig lang tid. KABI lager profesjonelle rapport-maler med automatisk datafylling – alle klientrapporter i ett konsistent format.",
    90:  "Hei Kristiansen Jus! Kontraktutkast som startes fra bunnen hver gang er en enorm tidtyv. Vi lager en kontraktmal-generator som produserer skreddersydde utkast automatisk basert på klienttype og sakstype.",
    91:  "Hei Fjord Regnskap & Skatt! MVA-rapportering som kompileres manuelt øker risikoen for feil og forsinkelser. KABI lager et MVA-rapport Excel som automatisk sammenstiller alle nødvendige data fra klientregnskapene.",
    92:  "Hei Helgeland Revisjon! Manuell intern rapportering til styret er tidkrevende og vanskelig å standardisere. Vi lager en styresrapport-generator som produserer profesjonelle styredokumenter automatisk fra regnskapsdata.",
    93:  "Hei Bymann Advokat! Fakturaer laget i Word er uprofesjonelle og vanskelige å spore. KABI lager en faktura-generator i Excel som produserer konsistente, profesjonelle fakturaer og holder oversikt over alle utestående.",
    94:  "Hei Østnorsk Regnskap! Manuell månedlig lønnsoversikt til klienter er tidkrevende og repetitivt arbeid. Vi bygger en lønnsrapport-generator som automatisk produserer skreddersydde lønnsoversikter per klient.",
    95:  "Hei Juris Finans Oslo! Klientoppfølging som mangler etter prosjektslutt gir tapte muligheter for mersalg. KABI lager en automatisert oppfølgingssekvens med e-postmaler og tracker som holder relasjonen varm.",
    96:  "Hei Hordaland Regnskapsservice! Manuell innsamling av regnskapsdata fra flere systemer er tidkrevende og feilutsatt. Vi bygger et Excel-aggregeringsverktøy som automatisk konsoliderer data fra alle kildene.",
    97:  "Hei Lunde Advokatkontor! Ufullstendige og rotete saksnotater øker risikoen for feil og setter profesjonaliteten på spill. KABI lager et notatsystem med standardiserte maler og PDF-ekstraktor for strukturert arkivering.",
    98:  "Hei Finans & Lov! Uten strukturert klientonboarding bruker dere unødvendig tid på å komme i gang med nye klienter. Vi lager en onboarding-mal og tracker som sikrer at alle nye klienter behandles konsistent og effektivt.",
    99:  "Hei Fjellstad Regnskap! Manglende automatisk varsling om frister er en forretningsrisiko. KABI lager en frist-tracker med automatiske e-postvarsler som sørger for at ingen viktige datoer noen gang glemmes.",
    100: "Hei Solberg & Partnere Advokater! Manuelle timerapporter til klienter er tidkrevende og gir rom for feil. Vi lager en timerapport-generator som automatisk sammenstiller fakturerbare timer og produserer profesjonelle rapporter.",
}

categories = [
    ("Håndverkere", list(range(1, 31))),
    ("Frilansere og konsulenter", list(range(31, 56))),
    ("Nettbutikker", list(range(56, 81))),
    ("Regnskapsførere og advokater", list(range(81, 101))),
]

lines = []
lines.append("# KABI AUTOMATION — Outreach-liste")
lines.append("")
lines.append("100 potensielle norske småbedriftskunder med personaliserte oppsøkende meldinger.")
lines.append("")
lines.append("---")
lines.append("")

kunde_by_nr = {k["Nr"]: k for k in kunder}

for cat_name, nr_range in categories:
    lines.append(f"## {cat_name} ({len(nr_range)} kunder)")
    lines.append("")
    for nr in nr_range:
        k = kunde_by_nr[nr]
        msg = outreach_messages[nr]
        lines.append(f"### {nr}. {k['Bedriftsnavn']} — {k['By']}")
        lines.append(f"**Problem:** {k['Manuelt problem']}")
        lines.append(f"**Melding:**")
        lines.append(f"> {msg}")
        lines.append("")
    lines.append("---")
    lines.append("")

with open(MD_PATH, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

print(f"  Lagret: {MD_PATH}")

# ── 4. SUMMARY ────────────────────────────────────────────────────────────────

print("")
print("=" * 55)
print("  FERDIG — KABI AUTOMATION KUNDELISTE")
print("=" * 55)
print(f"  Totalt kunder:         {len(kunder)}")
print(f"  Haandverkere:          {len([k for k in kunder if k['Nr'] <= 30])}")
print(f"  Frilansere/konsulenter:{len([k for k in kunder if 31 <= k['Nr'] <= 55])}")
print(f"  Nettbutikker:          {len([k for k in kunder if 56 <= k['Nr'] <= 80])}")
print(f"  Regnskap/advokater:    {len([k for k in kunder if k['Nr'] >= 81])}")
print(f"  Excel-fil:             kundeliste.xlsx")
print(f"  Outreach-fil:          outreach-liste.md")
print("=" * 55)
