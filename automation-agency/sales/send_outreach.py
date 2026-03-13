"""
send_outreach.py — KABI AUTOMATION Outreach-skript
====================================================
Leser kundeliste.xlsx og outreach-liste.md, åpner Gmail i nettleseren
med ferdig utfylt melding for hver kunde, og sporer status i Excel.

Kjør:  python send_outreach.py
Krev:  pandas, openpyxl
"""

import os
import re
import webbrowser
import urllib.parse
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ── Filstier ─────────────────────────────────────────────────────────────────

SCRIPT_DIR   = Path(__file__).parent
EXCEL_FIL    = SCRIPT_DIR / "kundeliste.xlsx"
OUTREACH_FIL = SCRIPT_DIR / "outreach-liste.md"

# ── Gmail URL-bygging ─────────────────────────────────────────────────────────

GMAIL_COMPOSE = "https://mail.google.com/mail/?view=cm&fs=1"

def bygg_gmail_url(emne: str, tekst: str) -> str:
    """Returnerer en Gmail compose-URL med ferdig utfylt emne og brødtekst."""
    params = urllib.parse.urlencode({
        "su": emne,
        "body": tekst,
    }, quote_via=urllib.parse.quote)
    return f"{GMAIL_COMPOSE}&{params}"

# ── Parser outreach-liste.md ──────────────────────────────────────────────────

def les_meldinger(fil: Path) -> dict[int, dict]:
    """
    Parser outreach-liste.md og returnerer:
        {nr: {"bedrift": str, "by": str, "melding": str}}
    """
    resultat = {}
    try:
        tekst = fil.read_text(encoding="utf-8")
    except FileNotFoundError:
        print(f"[FEIL] Finner ikke {fil}")
        return {}

    # Splitt på "### " for å hente hver kundeblokk
    blokker = re.split(r"\n### ", "\n" + tekst)
    for blokk in blokker:
        # Hopp over intro-tekst
        hode_match = re.match(r"(\d+)\.\s+(.+?)\s+—\s+(.+?)[\r\n]", blokk)
        if not hode_match:
            continue
        nr      = int(hode_match.group(1))
        bedrift = hode_match.group(2).strip()
        by      = hode_match.group(3).strip()

        # Hent teksten etter "> " (sitatlinjene)
        sitat_linjer = re.findall(r"^> (.+)", blokk, re.MULTILINE)
        melding = " ".join(sitat_linjer).strip()

        resultat[nr] = {
            "bedrift": bedrift,
            "by":      by,
            "melding": melding,
        }

    return resultat

# ── Excel-håndtering ──────────────────────────────────────────────────────────

def les_excel(fil: Path) -> pd.DataFrame:
    """Leser kundeliste og legger til sporings-kolonner om de mangler."""
    df = pd.read_excel(fil)

    if "Status" not in df.columns:
        df["Status"] = ""
    if "Kontaktet dato" not in df.columns:
        df["Kontaktet dato"] = ""

    return df

def lagre_excel(df: pd.DataFrame, fil: Path) -> None:
    """Lagrer DataFrame tilbake til Excel med formatering."""
    with pd.ExcelWriter(fil, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Kunder")
        ws = writer.sheets["Kunder"]

        # Kolonnebredder
        bredder = {
            "A": 5,   # Nr
            "B": 30,  # Bedriftsnavn
            "C": 20,  # Bransje
            "D": 15,  # By
            "E": 40,  # Manuelt problem
            "F": 30,  # KABI-løsning
            "G": 14,  # Pris (NOK)
            "H": 22,  # Finn dem på
            "I": 18,  # Status
            "J": 16,  # Kontaktet dato
        }
        for kol, bredde in bredder.items():
            ws.column_dimensions[kol].width = bredde

        # Topplinje-formatering
        topplinje_fyll = PatternFill("solid", fgColor="111111")
        topplinje_font = Font(bold=True, color="FFFFFF", size=11)
        for celle in ws[1]:
            celle.fill = topplinje_fyll
            celle.font = topplinje_font
            celle.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        # Status-farger
        grønn = PatternFill("solid", fgColor="C6EFCE")
        gul   = PatternFill("solid", fgColor="FFEB9C")
        for rad in ws.iter_rows(min_row=2):
            status_celle = rad[8]  # Kolonne I = Status
            if status_celle.value == "Sendt":
                status_celle.fill = grønn
            elif status_celle.value == "Åpnet":
                status_celle.fill = gul

    print(f"  OK Excel lagret: {fil.name}")

# ── Emnelinjer per bransje ─────────────────────────────────────────────────────

EMNER = {
    "Rørlegger":         "Spar timer hver uke – automatiser det manuelle arbeidet",
    "Elektriker":        "Automatiser timelistene og faktureringen din",
    "Snekker":           "Send profesjonelle tilbud på minutter, ikke timer",
    "Maler":             "Spar tid på rapporter – jobb med det du er god på",
    "Tømrer":            "Full kontroll over jobber og materialer – automatisk",
    "Frilans":           "Automatiser rapporter og fakturaer – og fakturér mer",
    "Konsulent":         "Klientrapporter ferdig på minutter – ikke timer",
    "Regnskap":          "Automatiser månedlige rapporter til alle klienter",
    "Advokat":           "Spar fakturerbar tid på dokumentsortering og maler",
    "Nettbutikk":        "Spar tid på bestillinger, lager og rapportering",
    "Revisjon":          "Automatiser årsoppgjør og klientrapporter",
    "Markedsføring":     "Rapporter ferdig på minutter – mer tid til klienter",
    "Design":            "Profesjonell fakturering og tilbud – automatisk",
    "Rekruttering":      "Automatiser kandidatoversikt og oppfølging",
}

def hent_emne(bransje: str) -> str:
    for nøkkel, emne in EMNER.items():
        if nøkkel.lower() in str(bransje).lower():
            return emne
    return "KABI Automation – Spar tid og jobb smartere"

# ── Signatur-tekst som legges til meldingen ───────────────────────────────────

SIGNATUR = """

---
Med vennlig hilsen,

Karlo Bikic
Grunnlegger, KABI AUTOMATION
Vi automatiserer det du gjør manuelt

📧 karlo@kabiautomation.com
📱 [ditt telefonnummer]

PS: Svarer du på denne e-posten, setter jeg opp et gratis 20-minutters møte
for å vise deg nøyaktig hva vi kan automatisere for deg.
"""

# ── Statistikk-visning ────────────────────────────────────────────────────────

def vis_statistikk(df: pd.DataFrame) -> None:
    totalt     = len(df)
    sendt      = (df["Status"] == "Sendt").sum()
    ikke_sendt = totalt - sendt
    print(f"\n{'-'*50}")
    print(f"  STATUS-OVERSIKT")
    print(f"{'-'*50}")
    print(f"  Totalt antall kunder : {totalt}")
    print(f"  Sendt outreach       : {sendt}")
    print(f"  Ikke kontaktet       : {ikke_sendt}")
    if sendt > 0:
        print(f"\n  Siste 5 kontaktede:")
        siste = df[df["Status"] == "Sendt"].tail(5)
        for _, rad in siste.iterrows():
            print(f"    #{int(rad['Nr']):>3}  {rad['Bedriftsnavn']:<30}  {rad['Kontaktet dato']}")
    print(f"{'-'*50}\n")

# ── Hoved-logikk ──────────────────────────────────────────────────────────────

def main():
    print("\n" + "="*55)
    print("  KABI AUTOMATION - Outreach-skript")
    print("="*55)

    # Last inn data
    if not EXCEL_FIL.exists():
        print(f"[FEIL] Finner ikke {EXCEL_FIL}")
        return
    if not OUTREACH_FIL.exists():
        print(f"[FEIL] Finner ikke {OUTREACH_FIL}")
        return

    df       = les_excel(EXCEL_FIL)
    meldinger = les_meldinger(OUTREACH_FIL)
    print(f"  OK Lastet {len(df)} kunder fra Excel")
    print(f"  OK Lastet {len(meldinger)} outreach-meldinger\n")

    # Vis statistikk
    vis_statistikk(df)

    # Meny
    print("  VELG HANDLING:")
    print("  [1]  Send outreach til de neste N ikke-kontaktede kundene")
    print("  [2]  Send til spesifikk kundrekke (f.eks. 1-10)")
    print("  [3]  Send til en spesifikk kunde (kundnr)")
    print("  [4]  Vis statistikk og avslutt")
    print()

    valg = input("  Ditt valg [1-4]: ").strip()
    print()

    if valg == "4":
        vis_statistikk(df)
        return

    # Bestem hvilke kunder som skal prosesseres
    mål_kunder = []

    if valg == "1":
        antall = input("  Antall kunder å sende til (f.eks. 10): ").strip()
        try:
            antall = int(antall)
        except ValueError:
            print("[FEIL] Ugyldig antall")
            return
        ikke_kontaktet = df[df["Status"] != "Sendt"]
        mål_kunder = ikke_kontaktet.head(antall).index.tolist()

    elif valg == "2":
        rekke = input("  Kundrekke (f.eks. 1-10): ").strip()
        try:
            start, slutt = rekke.split("-")
            start, slutt = int(start), int(slutt)
        except Exception:
            print("[FEIL] Ugyldig format. Bruk f.eks. 1-10")
            return
        mål_kunder = df[(df["Nr"] >= start) & (df["Nr"] <= slutt)].index.tolist()

    elif valg == "3":
        nr = input("  Kundenummer: ").strip()
        try:
            nr = int(nr)
        except ValueError:
            print("[FEIL] Ugyldig kundenummer")
            return
        mål_kunder = df[df["Nr"] == nr].index.tolist()
        if not mål_kunder:
            print(f"[FEIL] Finner ikke kunde #{nr}")
            return

    else:
        print("[FEIL] Ugyldig valg")
        return

    if not mål_kunder:
        print("  Ingen kunder å prosessere.")
        return

    print(f"\n  Klar til å prosessere {len(mål_kunder)} kunde(r).\n")
    print("  Trykk ENTER for å åpne Gmail for hver kunde.")
    print("  Skriv 'hopp' for å hoppe over, 'ferdig' for å avslutte.\n")

    endret = 0
    for idx in mål_kunder:
        rad     = df.loc[idx]
        nr      = int(rad["Nr"])
        bedrift = rad["Bedriftsnavn"]
        bransje = rad.get("Bransje", "")
        by      = rad.get("By", "")

        # Hent melding
        melding_data = meldinger.get(nr, {})
        melding_tekst = melding_data.get("melding", f"Hei {bedrift}! KABI Automation hjelper deg med å automatisere manuelle prosesser og spare tid.")

        full_melding = melding_tekst + SIGNATUR
        emne = hent_emne(bransje)

        print(f"  Kunde #{nr:>3}: {bedrift} ({by})")
        print(f"  Bransje : {bransje}")
        print(f"  Emne    : {emne}")
        print(f"  Melding : {melding_tekst[:80]}...")
        print()

        svar = input("  [ENTER = åpne Gmail | 'hopp' = hopp over | 'ferdig' = avslutt]: ").strip().lower()

        if svar == "ferdig":
            print("\n  Avslutter...")
            break
        elif svar == "hopp":
            print(f"  >> Hoppet over #{nr}\n")
            continue
        else:
            # Åpne Gmail
            url = bygg_gmail_url(emne, full_melding)
            webbrowser.open(url)
            print(f"  OK Gmail aapnet for {bedrift}")

            # Oppdater status
            df.at[idx, "Status"]          = "Sendt"
            df.at[idx, "Kontaktet dato"]  = date.today().strftime("%d.%m.%Y")
            endret += 1
            print(f"  OK Status oppdatert: Sendt ({date.today().strftime('%d.%m.%Y')})\n")

    # Lagre
    if endret > 0:
        lagre_excel(df, EXCEL_FIL)
        print(f"\n  OK {endret} kunde(r) oppdatert og lagret i Excel.\n")
    else:
        print("\n  Ingen endringer lagret.\n")

    vis_statistikk(df)


if __name__ == "__main__":
    main()
