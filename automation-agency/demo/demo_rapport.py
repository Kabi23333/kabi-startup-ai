"""
demo_rapport.py — Imponerende demo-rapport for KABI AUTOMATION
Bruk: python demo_rapport.py testdata_ror_as.xlsx
      (eller: python demo_rapport.py [hvilken som helst Excel-fil])
Krever: pandas, openpyxl, fpdf2, anthropic
"""

import sys
import os
import pandas as pd
from datetime import datetime
import anthropic

try:
    from fpdf import FPDF
except ImportError:
    print("FEIL: 'fpdf2' er ikke installert. Kjoer: pip install fpdf2")
    sys.exit(1)


# ─────────────────────────────────────────────
# DATAHENTING OG BEHANDLING
# ─────────────────────────────────────────────

def les_og_rens(filsti):
    """Les Excel-fil og rens data."""
    print(f"  Leser {filsti}...")
    df = pd.read_excel(filsti)

    # Normaliser kolonnenavn
    df.columns = [str(c).strip() for c in df.columns]

    # Fjern tomme rader
    df = df.dropna(how="all").reset_index(drop=True)

    # Rens whitespace
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].str.strip()

    # Konverter Belop til numerisk
    if "Belop" in df.columns:
        df["Belop"] = pd.to_numeric(df["Belop"], errors="coerce")

    # Konverter Dato og legg til Maned
    if "Dato" in df.columns:
        df["Dato"] = pd.to_datetime(df["Dato"], dayfirst=True, errors="coerce")
        df["Maned"] = df["Dato"].dt.strftime("%Y-%m")
        df["ManedNavn"] = df["Dato"].dt.strftime("%B %Y")

    print(f"  {len(df)} rader lest og renset")
    return df


def beregn_nokkeltal(df):
    """Beregn alle nokkeltal."""
    belop_col = "Belop"

    inntekter_df = df[df[belop_col] > 0].copy()
    utgifter_df = df[df[belop_col] < 0].copy()

    total_inntekt = inntekter_df[belop_col].sum()
    total_utgift = utgifter_df[belop_col].sum()
    resultat = total_inntekt + total_utgift
    margin = (resultat / total_inntekt * 100) if total_inntekt != 0 else 0

    # Beste og verste maned
    if "Maned" in df.columns:
        maned_resultat = df.groupby("Maned")[belop_col].sum()
        beste_maned = maned_resultat.idxmax()
        verste_maned = maned_resultat.idxmin()
        beste_belop = maned_resultat[beste_maned]
        verste_belop = maned_resultat[verste_maned]
    else:
        beste_maned = verste_maned = "Ukjent"
        beste_belop = verste_belop = 0

    # Topp inntektskategori
    if "Kategori" in df.columns:
        inntekt_per_kat = inntekter_df.groupby("Kategori")[belop_col].sum()
        topp_inntekt_kat = inntekt_per_kat.idxmax() if not inntekt_per_kat.empty else "Ukjent"
        topp_inntekt_belop = inntekt_per_kat.max() if not inntekt_per_kat.empty else 0

        utgift_per_kat = utgifter_df.groupby("Kategori")[belop_col].sum()
        storste_utgift_kat = utgift_per_kat.idxmin() if not utgift_per_kat.empty else "Ukjent"
        storste_utgift_belop = utgift_per_kat.min() if not utgift_per_kat.empty else 0
    else:
        topp_inntekt_kat = storste_utgift_kat = "Ukjent"
        topp_inntekt_belop = storste_utgift_belop = 0

    return {
        "total_inntekt": total_inntekt,
        "total_utgift": total_utgift,
        "resultat": resultat,
        "margin": margin,
        "beste_maned": beste_maned,
        "beste_belop": beste_belop,
        "verste_maned": verste_maned,
        "verste_belop": verste_belop,
        "topp_inntekt_kat": topp_inntekt_kat,
        "topp_inntekt_belop": topp_inntekt_belop,
        "storste_utgift_kat": storste_utgift_kat,
        "storste_utgift_belop": storste_utgift_belop,
        "antall_transaksjoner": len(df),
        "antall_inntekter": len(inntekter_df),
        "antall_utgifter": len(utgifter_df),
    }


# ─────────────────────────────────────────────
# AI-SAMMENDRAG
# ─────────────────────────────────────────────

def generer_ai_sammendrag(nokkeltal, filnavn):
    """Bruk Claude til a generere et norsk forretningssammendrag."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return (
            "AI-sammendrag ikke tilgjengelig (ANTHROPIC_API_KEY ikke satt). "
            "Kontakt KABI AUTOMATION for full rapport."
        )

    klient = anthropic.Anthropic(api_key=api_key)

    prompt = f"""Du er en okonomisk radgiver for norske smaabedrifter.
Skriv et profesjonelt og handlingsorientert sammendrag pa 5-7 setninger basert pa disse tallene:

Bedrift/fil: {filnavn}
Totale inntekter: {nokkeltal['total_inntekt']:,.0f} NOK
Totale utgifter: {nokkeltal['total_utgift']:,.0f} NOK
Resultat: {nokkeltal['resultat']:,.0f} NOK
Margin: {nokkeltal['margin']:.1f}%
Beste maned: {nokkeltal['beste_maned']} ({nokkeltal['beste_belop']:,.0f} NOK nettoresultat)
Verste maned: {nokkeltal['verste_maned']} ({nokkeltal['verste_belop']:,.0f} NOK nettoresultat)
Storste inntektskategori: {nokkeltal['topp_inntekt_kat']} ({nokkeltal['topp_inntekt_belop']:,.0f} NOK)
Storste utgiftskategori: {nokkeltal['storste_utgift_kat']} ({abs(nokkeltal['storste_utgift_belop']):,.0f} NOK)

Inkluder konkrete observasjoner og 1-2 handlingsrettede rad. Skriv direkte til bedriftseieren. Unnga teknisk sjargong."""

    try:
        melding = klient.messages.create(
            model="claude-opus-4-5",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}]
        )
        return melding.content[0].text
    except anthropic.AuthenticationError:
        return "AI-sammendrag utilgjengelig — ugyldig API-nokkel."
    except Exception as e:
        return f"AI-sammendrag utilgjengelig: {e}"


# ─────────────────────────────────────────────
# EXCEL-RAPPORT
# ─────────────────────────────────────────────

def lagre_excel_rapport(df, nokkeltal, ai_sammendrag, utfil):
    """Lagre formatert Excel-rapport med tre ark."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    print(f"  Lagrer Excel: {utfil}")

    with pd.ExcelWriter(utfil, engine="openpyxl") as writer:

        # ── ARK 1: Sammendrag ──
        sammendrag_data = {
            "Nokkeltal": [
                "Totale inntekter",
                "Totale utgifter",
                "Nettoresultat",
                "Margin",
                "Antall transaksjoner",
                "Antall inntekter",
                "Antall utgifter",
                "Beste maned",
                "Verste maned",
                "Storste inntektskategori",
                "Storste utgiftskategori",
            ],
            "Verdi": [
                f"{nokkeltal['total_inntekt']:,.0f} NOK",
                f"{nokkeltal['total_utgift']:,.0f} NOK",
                f"{nokkeltal['resultat']:,.0f} NOK",
                f"{nokkeltal['margin']:.1f}%",
                str(nokkeltal["antall_transaksjoner"]),
                str(nokkeltal["antall_inntekter"]),
                str(nokkeltal["antall_utgifter"]),
                f"{nokkeltal['beste_maned']} ({nokkeltal['beste_belop']:,.0f} NOK)",
                f"{nokkeltal['verste_maned']} ({nokkeltal['verste_belop']:,.0f} NOK)",
                f"{nokkeltal['topp_inntekt_kat']} ({nokkeltal['topp_inntekt_belop']:,.0f} NOK)",
                f"{nokkeltal['storste_utgift_kat']} ({abs(nokkeltal['storste_utgift_belop']):,.0f} NOK)",
            ]
        }
        df_sammendrag = pd.DataFrame(sammendrag_data)
        df_sammendrag.to_excel(writer, index=False, sheet_name="Sammendrag", startrow=3)

        ws = writer.sheets["Sammendrag"]

        # Tittel
        ws["A1"] = "KABI AUTOMATION — Automatisk Okonomisk Rapport"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:B1")

        ws["A2"] = f"Generert: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=10, color="666666")
        ws.merge_cells("A2:B2")

        # Formatering av kolonner
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 40

        # Fargelegg header-rad
        for cell in ws[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")

        # Fargelegg annenhver rad
        for row_idx in range(5, 5 + len(df_sammendrag)):
            fill_color = "DEEAF1" if (row_idx % 2 == 0) else "FFFFFF"
            for cell in ws[row_idx]:
                cell.fill = PatternFill("solid", fgColor=fill_color)

        # AI-sammendrag
        ai_rad = 5 + len(df_sammendrag) + 2
        ws.cell(row=ai_rad, column=1, value="AI-ANALYSE")
        ws.cell(row=ai_rad, column=1).font = Font(bold=True, size=11, color="1F4E79")
        ws.cell(row=ai_rad + 1, column=1, value=ai_sammendrag)
        ws.cell(row=ai_rad + 1, column=1).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[ai_rad + 1].height = 120
        ws.merge_cells(f"A{ai_rad + 1}:B{ai_rad + 1}")

        # ── ARK 2: Transaksjoner ──
        df_vis = df.drop(columns=["Maned", "ManedNavn"], errors="ignore").copy()
        if "Dato" in df_vis.columns:
            df_vis["Dato"] = df_vis["Dato"].dt.strftime("%d.%m.%Y")
        df_vis.to_excel(writer, index=False, sheet_name="Transaksjoner")

        ws2 = writer.sheets["Transaksjoner"]
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
        ws2.column_dimensions["A"].width = 14
        ws2.column_dimensions["C"].width = 28
        ws2.column_dimensions["D"].width = 38
        ws2.column_dimensions["E"].width = 22

        # ── ARK 3: Kategori ──
        if "Kategori" in df.columns and "Belop" in df.columns:
            kat_df = df.groupby("Kategori")["Belop"].agg(
                Totalt="sum", Antall="count", Gjennomsnitt="mean"
            ).reset_index()
            kat_df.columns = ["Kategori", "Totalt (NOK)", "Antall", "Snitt (NOK)"]
            kat_df = kat_df.sort_values("Totalt (NOK)", ascending=False)
            kat_df.to_excel(writer, index=False, sheet_name="Kategori")

            ws3 = writer.sheets["Kategori"]
            for cell in ws3[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
            ws3.column_dimensions["A"].width = 28

    print("  Excel-rapport lagret!")


# ─────────────────────────────────────────────
# PDF-RAPPORT
# ─────────────────────────────────────────────

class RapportPDF(FPDF):
    def __init__(self):
        super().__init__()
        self.add_font("Arial", "", "C:\\Windows\\Fonts\\arial.ttf")
        self.add_font("Arial", "B", "C:\\Windows\\Fonts\\arialbd.ttf")
        self.add_font("Arial", "I", "C:\\Windows\\Fonts\\ariali.ttf")

    def header(self):
        self.set_fill_color(31, 78, 121)
        self.rect(0, 0, 210, 20, "F")
        self.set_font("Arial", "B", 14)
        self.set_text_color(255, 255, 255)
        self.set_xy(10, 5)
        self.cell(0, 10, "KABI AUTOMATION   Automatisk Økonomisk Rapport", align="L")
        self.set_text_color(0, 0, 0)
        self.ln(18)

    def footer(self):
        self.set_y(-12)
        self.set_font("Arial", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 5, f"Generert av KABI AUTOMATION  |  {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  Side {self.page_no()}", align="C")


def lagre_pdf_rapport(nokkeltal, ai_sammendrag, df, utfil):
    """Lag en profesjonell PDF-rapport."""
    print(f"  Lagrer PDF: {utfil}")

    pdf = RapportPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ── Dato og fil ──
    pdf.set_font("Arial", "I", 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, f"Rapport generert: {datetime.now().strftime('%d.%m.%Y kl. %H:%M')}", ln=True)
    pdf.ln(4)

    # ── Nøkkeltall-bokser ──
    pdf.set_font("Arial", "B", 11)
    pdf.set_text_color(31, 78, 121)
    pdf.cell(0, 8, "NØKKELTALL", ln=True)
    pdf.ln(2)

    def boks(tittel, verdi, farge_rgb, x, y, bredde=58, hoyde=22):
        pdf.set_xy(x, y)
        pdf.set_fill_color(*farge_rgb)
        pdf.rect(x, y, bredde, hoyde, "F")
        pdf.set_xy(x + 2, y + 2)
        pdf.set_font("Arial", "", 8)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(bredde - 4, 5, tittel, ln=True)
        pdf.set_xy(x + 2, y + 8)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(bredde - 4, 9, verdi, ln=True)

    y_start = pdf.get_y()
    boks("TOTALE INNTEKTER", f"{nokkeltal['total_inntekt']:,.0f} NOK", (46, 125, 50), 10, y_start)
    boks("TOTALE UTGIFTER", f"{abs(nokkeltal['total_utgift']):,.0f} NOK", (198, 40, 40), 72, y_start)
    boks("NETTORESULTAT", f"{nokkeltal['resultat']:,.0f} NOK",
         (31, 78, 121) if nokkeltal['resultat'] >= 0 else (150, 30, 30), 134, y_start)

    pdf.set_y(y_start + 28)
    y_start2 = pdf.get_y()
    boks("MARGIN", f"{nokkeltal['margin']:.1f}%", (87, 87, 87), 10, y_start2)
    boks("BESTE MÅNED", nokkeltal["beste_maned"], (46, 125, 50), 72, y_start2)
    boks("VERSTE MÅNED", nokkeltal["verste_maned"], (198, 40, 40), 134, y_start2)

    pdf.set_y(y_start2 + 28)
    pdf.ln(4)

    # ── AI-analyse ──
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "B", 11)
    pdf.set_text_color(31, 78, 121)
    pdf.cell(0, 8, "AI-ANALYSE", ln=True)

    pdf.set_fill_color(222, 234, 241)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 9)

    # Strip markdown-formatering fra Claude-output
    def rens_ai_tekst(t):
        t = t.replace("\r", "")
        t = t.replace("**", "").replace("## ", "").replace("# ", "")
        return t

    tekstlinjer = rens_ai_tekst(ai_sammendrag)
    x_marg = 10
    bredde = 190

    pdf.set_xy(x_marg, pdf.get_y())
    pdf.set_fill_color(222, 234, 241)
    pdf.multi_cell(bredde, 5, tekstlinjer, fill=True)
    pdf.ln(6)

    # ── Kategorioversikt ──
    if "Kategori" in df.columns and "Belop" in df.columns:
        pdf.set_font("Arial", "B", 11)
        pdf.set_text_color(31, 78, 121)
        pdf.cell(0, 8, "KATEGORIOVERSIKT", ln=True)
        pdf.ln(2)

        kat_df = df.groupby("Kategori")["Belop"].sum().reset_index()
        kat_df.columns = ["Kategori", "Totalt"]
        kat_df = kat_df.sort_values("Totalt", ascending=False)

        # Tabellheader
        col_w = [90, 50, 40]
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 9)
        pdf.cell(col_w[0], 7, "Kategori", border=0, fill=True)
        pdf.cell(col_w[1], 7, "Totalt (NOK)", border=0, fill=True, align="R")
        pdf.cell(col_w[2], 7, "Type", border=0, fill=True, align="C")
        pdf.ln()

        pdf.set_font("Arial", "", 9)
        for i, (_, rad) in enumerate(kat_df.iterrows()):
            fill_color = (222, 234, 241) if i % 2 == 0 else (255, 255, 255)
            pdf.set_fill_color(*fill_color)
            pdf.set_text_color(0, 0, 0)
            type_label = "Inntekt" if rad["Totalt"] > 0 else "Utgift"
            pdf.cell(col_w[0], 6, str(rad["Kategori"]), border=0, fill=True)
            pdf.cell(col_w[1], 6, f"{rad['Totalt']:,.0f}", border=0, fill=True, align="R")
            pdf.cell(col_w[2], 6, type_label, border=0, fill=True, align="C")
            pdf.ln()

    pdf.output(utfil)
    print("  PDF-rapport lagret!")


# ─────────────────────────────────────────────
# HOVEDFUNKSJON
# ─────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        innfil = "testdata_ror_as.xlsx"
        print(f"Ingen fil oppgitt — bruker standard: {innfil}")
    else:
        innfil = sys.argv[1]

    if not os.path.exists(innfil):
        print(f"FEIL: Filen '{innfil}' ble ikke funnet.")
        print("Tips: Kjoer 'python lag_demo_data.py' forst for a lage testdata.")
        sys.exit(1)

    base_navn = os.path.splitext(os.path.basename(innfil))[0]
    utfil_excel = f"RAPPORT_{base_navn}.xlsx"
    utfil_pdf = f"RAPPORT_{base_navn}.pdf"

    print("\n" + "=" * 60)
    print("  KABI AUTOMATION — Automatisk Okonomisk Analyse")
    print("=" * 60)
    print(f"\nAnalyserer: {innfil}")
    print("Dette vil ta noen sekunder...\n")

    # Steg 1: Les og rens data
    print("[1/5] Leser og renser data...")
    df = les_og_rens(innfil)

    # Steg 2: Beregn nokkeltal
    print("\n[2/5] Beregner nokkeltal...")
    nokkeltal = beregn_nokkeltal(df)

    print(f"\n  Totale inntekter:  {nokkeltal['total_inntekt']:>12,.0f} NOK")
    print(f"  Totale utgifter:   {nokkeltal['total_utgift']:>12,.0f} NOK")
    print(f"  Nettoresultat:     {nokkeltal['resultat']:>12,.0f} NOK")
    print(f"  Margin:            {nokkeltal['margin']:>11.1f}%")
    print(f"  Beste maned:       {nokkeltal['beste_maned']} ({nokkeltal['beste_belop']:,.0f} NOK)")
    print(f"  Verste maned:      {nokkeltal['verste_maned']} ({nokkeltal['verste_belop']:,.0f} NOK)")
    print(f"  Topp-inntektkat:   {nokkeltal['topp_inntekt_kat']}")
    print(f"  Storste utgift:    {nokkeltal['storste_utgift_kat']}")

    # Steg 3: AI-sammendrag
    print("\n[3/5] Genererer AI-sammendrag (Claude)...")
    ai_sammendrag = generer_ai_sammendrag(nokkeltal, base_navn)
    print("\n  AI sier:")
    print("  " + ai_sammendrag.replace("\n", "\n  "))

    # Steg 4: Excel-rapport
    print(f"\n[4/5] Lagrer Excel-rapport...")
    lagre_excel_rapport(df, nokkeltal, ai_sammendrag, utfil_excel)

    # Steg 5: PDF-rapport
    print(f"\n[5/5] Lagrer PDF-rapport...")
    lagre_pdf_rapport(nokkeltal, ai_sammendrag, df, utfil_pdf)

    # Ferdig
    print("\n" + "=" * 60)
    print(f"Rapporten er klar! Sjekk {utfil_excel} og {utfil_pdf}")
    print("=" * 60)
    print(f"\nFiler generert:")
    print(f"  Excel: {utfil_excel}")
    print(f"  PDF:   {utfil_pdf}")


if __name__ == "__main__":
    main()
